from rest_framework import status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
from django.http import HttpResponse
import csv
import io
import openpyxl

from apps.contacts.models import ContactGroup, Contact, ContactGroupMembership
from apps.contacts.serializers import (
    ContactGroupSerializer, ContactSerializer, ContactGroupMembershipSerializer,
    ContactImportSerializer
)
from apps.common.viewsets import BaseOrganizationViewSet
from apps.common.tasks import import_contacts_from_csv
from apps.subscriptions.subscription_views import check_usage_limit, update_usage_tracking


class DefaultPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


class ContactGroupViewSet(BaseOrganizationViewSet):
    queryset = ContactGroup.objects.all()
    serializer_class = ContactGroupSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['name', 'description']

    @action(detail=True, methods=['post'])
    def add_contacts(self, request, pk=None):
        """Add contacts to group"""
        group = self.get_object()
        contact_ids = request.data.get('contact_ids', [])
        
        contacts = Contact.objects.filter(
            id__in=contact_ids,
            organization=self.request.user.organization
        )
        
        for contact in contacts:
            ContactGroupMembership.objects.get_or_create(
                contact=contact,
                group=group
            )
        
        return Response({'message': f'Added {contacts.count()} contacts to group'})

    @action(detail=True, methods=['post'])
    def remove_contacts(self, request, pk=None):
        """Remove contacts from group"""
        group = self.get_object()
        contact_ids = request.data.get('contact_ids', [])
        
        ContactGroupMembership.objects.filter(
            contact_id__in=contact_ids,
            group=group
        ).delete()
        
        return Response({'message': 'Contacts removed from group'})

    @action(detail=True, methods=['get'])
    def contacts(self, request, pk=None):
        """Get all contacts in this group"""
        group = self.get_object()
        
        # Get contacts that belong to this group
        contact_ids = ContactGroupMembership.objects.filter(
            group=group
        ).values_list('contact_id', flat=True)
        
        contacts = Contact.objects.filter(
            id__in=contact_ids,
            organization=self.request.user.organization
        )
        
        serializer = ContactSerializer(contacts, many=True)
        return Response(serializer.data)


class ContactViewSet(BaseOrganizationViewSet):
    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_subscribed', 'language']
    search_fields = ['email', 'first_name', 'last_name']
    pagination_class = DefaultPagination
    
    def create(self, request, *args, **kwargs):
        """Create a new contact with subscription limit check"""
        if not request.user.organization:
            return Response({
                'error': 'No organization found'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check contact limit
        limit_check = check_usage_limit(request.user.organization, 'contacts')
        if not limit_check['allowed']:
            return Response({
                'error': limit_check['reason'],
                'code': 'CONTACT_LIMIT_REACHED',
                'current': limit_check['current'],
                'limit': limit_check['limit'],
                'upgrade_url': '/subscription/upgrade'
            }, status=status.HTTP_403_FORBIDDEN)
        
        response = super().create(request, *args, **kwargs)
        
        # Update usage tracking
        if response.status_code == 201:
            update_usage_tracking(request.user.organization, 'contacts')
        
        return response

    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        """Import contacts from CSV file with subscription limit check"""
        # Check contact limit before import
        limit_check = check_usage_limit(request.user.organization, 'contacts')
        if not limit_check['allowed']:
            return Response({
                'error': limit_check['reason'],
                'code': 'CONTACT_LIMIT_REACHED',
                'current': limit_check['current'],
                'limit': limit_check['limit'],
                'upgrade_url': '/subscription/upgrade'
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = ContactImportSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            
            # Read file content
            if file.name.lower().endswith('.csv'):
                content = file.read().decode('utf-8')
            elif file.name.lower().endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                
                # Convert Excel to CSV format
                output = io.StringIO()
                writer = csv.writer(output)
                
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(row)
                
                content = output.getvalue()
            
            # Start background task
            task = import_contacts_from_csv.delay(
                self.request.user.organization.id,
                content,
                self.request.user.id
            )
            
            # Update usage tracking
            update_usage_tracking(request.user.organization, 'contacts')
            
            return Response({
                'message': 'Import started',
                'task_id': task.id
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """Bulk update contacts"""
        contact_ids = request.data.get('contact_ids', [])
        update_data = request.data.get('data', {})
        
        contacts = Contact.objects.filter(
            id__in=contact_ids,
            organization=self.request.user.organization
        )
        
        updated_count = contacts.update(**update_data)
        
        return Response({
            'message': f'Updated {updated_count} contacts'
        })

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export contacts as CSV or Excel file"""
        # Get format parameter (csv or xlsx)
        export_format = request.query_params.get('format', 'csv').lower()
        
        # Get filtered contacts based on query parameters
        queryset = self.get_queryset()
        
        # Apply filters similar to the list view
        search = request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(email__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search)
            )
        
        is_subscribed = request.query_params.get('is_subscribed', '')
        if is_subscribed in ['true', 'false']:
            queryset = queryset.filter(is_subscribed=is_subscribed == 'true')
        
        language = request.query_params.get('language', '')
        if language:
            queryset = queryset.filter(language=language)
        
        # Helper function to sanitize fields to prevent injection
        def sanitize_field(value):
            """Sanitize field to prevent CSV/Excel injection attacks"""
            if not value:
                return ''
            value_str = str(value)
            # Strip leading whitespace and tabs, then check for dangerous characters
            stripped_value = value_str.lstrip(' \t')
            if stripped_value.startswith(('=', '+', '-', '@')):
                return "'" + value_str
            return value_str
        
        # Prepare data rows
        headers = [
            'email', 'first_name', 'last_name', 'phone', 
            'language', 'is_subscribed', 'created_at', 'updated_at'
        ]
        
        rows = []
        for contact in queryset.order_by('email'):
            rows.append([
                sanitize_field(contact.email),
                sanitize_field(contact.first_name or ''),
                sanitize_field(contact.last_name or ''),
                sanitize_field(contact.phone or ''),
                sanitize_field(contact.language or 'en'),
                'Yes' if contact.is_subscribed else 'No',
                contact.created_at.strftime('%Y-%m-%d %H:%M:%S') if contact.created_at else '',
                contact.updated_at.strftime('%Y-%m-%d %H:%M:%S') if contact.updated_at else ''
            ])
        
        if export_format == 'xlsx':
            # Create Excel file
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Contacts"
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_idx, value=header)
            
            # Write data rows
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Create response
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="contacts_export.xlsx"'
            
        else:
            # Create CSV response (default)
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="contacts_export.csv"'
            
            writer = csv.writer(response)
            writer.writerow(headers)
            writer.writerows(rows)
        
        return response
