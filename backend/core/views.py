from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination


class DefaultPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Count, Sum
from django.db import transaction
from django.utils import timezone
import csv
import io
import openpyxl

from .models import (
    Organization, User, Domain, ContactGroup, Contact, 
    ContactGroupMembership, EmailTemplate, Campaign, 
    CampaignRecipient, AnalyticsEvent, Card
)
from .serializers import (
    OrganizationSerializer, UserSerializer, DomainSerializer,
    ContactGroupSerializer, ContactSerializer, ContactGroupMembershipSerializer,
    EmailTemplateSerializer, CampaignSerializer, CampaignRecipientSerializer,
    AnalyticsEventSerializer, DashboardStatsSerializer, ContactImportSerializer,
    CampaignSendSerializer, CardSerializer, CardCreateSerializer,
    CardUpdateSerializer
)
from .tasks import send_campaign_emails, verify_domain_dns, import_contacts_from_csv
from .subscription_views import (
    check_subscription_active, check_feature_available,
    check_usage_limit, get_current_usage, update_usage_tracking
)
from .middleware import requires_active_subscription, requires_plan_feature, track_usage


class BaseOrganizationViewSet(viewsets.ModelViewSet):
    """Base viewset that filters by organization"""
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.organization:
            return self.queryset.none()
        return self.queryset.filter(organization=self.request.user.organization)

    def perform_create(self, serializer):
        if self.request.user.organization:
            serializer.save(organization=self.request.user.organization)


class OrganizationViewSet(viewsets.ModelViewSet):
    queryset = Organization.objects.all()
    serializer_class = OrganizationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.user.is_superuser:
            return Organization.objects.all()
        if self.request.user.organization:
            return Organization.objects.filter(id=self.request.user.organization.id)
        return Organization.objects.none()


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if not self.request.user.organization:
            return User.objects.none()
        return User.objects.filter(organization=self.request.user.organization)

    @action(detail=False, methods=['get'])
    def me(self, request):
        """Get current user profile"""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

    @action(detail=False, methods=['patch'])
    def update_profile(self, request):
        """Update current user profile"""
        serializer = self.get_serializer(request.user, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class DomainViewSet(BaseOrganizationViewSet):
    queryset = Domain.objects.all()
    serializer_class = DomainSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['status']
    search_fields = ['domain']
    
    def create(self, request, *args, **kwargs):
        """Create a new domain with multi-domain check"""
        if not request.user.organization:
            return Response({
                'error': 'No organization found'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check domain limit
        limit_check = check_usage_limit(request.user.organization, 'domains')
        if not limit_check['allowed']:
            # Check if white labeling is available for unlimited domains
            if not check_feature_available(request.user.organization, 'has_white_labeling'):
                return Response({
                    'error': limit_check['reason'],
                    'code': 'DOMAIN_LIMIT_REACHED',
                    'current': limit_check['current'],
                    'limit': limit_check['limit'],
                    'note': 'Upgrade to Premium for unlimited custom domains',
                    'upgrade_url': '/subscription/upgrade'
                }, status=status.HTTP_403_FORBIDDEN)
        
        return super().create(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """Trigger domain verification with subscription check"""
        domain = self.get_object()
        
        # Check if subscription is active
        if not check_subscription_active(request.user.organization):
            return Response({
                'error': 'Active subscription required for domain verification',
                'code': 'SUBSCRIPTION_REQUIRED',
                'upgrade_url': '/subscription/plans'
            }, status=status.HTTP_402_PAYMENT_REQUIRED)
        
        verify_domain_dns.delay(domain.id)
        
        # Update usage tracking
        update_usage_tracking(request.user.organization, 'domains')
        
        return Response({'message': 'Domain verification started'})

    @action(detail=True, methods=['post'])
    def generate_dns_records(self, request, pk=None):
        """Generate DNS records for domain"""
        domain = self.get_object()
        
        # Generate DNS records (simplified)
        txt_record = f"engagex-verification={domain.id}"
        dkim_record = f"v=DKIM1; k=rsa; p=..."  # Would be actual DKIM key
        cname_record = "engagex.tracking.domain.com"
        dmarc_record = "v=DMARC1; p=none; rua=mailto:dmarc@engagex.com"
        
        domain.txt_record = txt_record
        domain.dkim_record = dkim_record
        domain.cname_record = cname_record
        domain.dmarc_record = dmarc_record
        domain.save()
        
        serializer = self.get_serializer(domain)
        return Response(serializer.data)


class ContactGroupViewSet(BaseOrganizationViewSet):
    queryset = ContactGroup.objects.all()
    serializer_class = ContactGroupSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    search_fields = ['name', 'description']

    @action(detail=True, methods=['post'])
    def add_contacts(self, request, pk=None):
        """Add contacts to group"""
        group = self.get_object()
        contact_ids = request.data.get('contact_ids', [])
        
        contacts = Contact.objects.filter(
            id__in=contact_ids,
            organization=self.request.user.organization
        )
        
        for contact in contacts:
            ContactGroupMembership.objects.get_or_create(
                contact=contact,
                group=group
            )
        
        return Response({'message': f'Added {contacts.count()} contacts to group'})

    @action(detail=True, methods=['post'])
    def remove_contacts(self, request, pk=None):
        """Remove contacts from group"""
        group = self.get_object()
        contact_ids = request.data.get('contact_ids', [])
        
        ContactGroupMembership.objects.filter(
            contact_id__in=contact_ids,
            group=group
        ).delete()
        
        return Response({'message': 'Contacts removed from group'})

    @action(detail=True, methods=['get'])
    def contacts(self, request, pk=None):
        """Get all contacts in this group"""
        group = self.get_object()
        
        # Get contacts that belong to this group
        contact_ids = ContactGroupMembership.objects.filter(
            group=group
        ).values_list('contact_id', flat=True)
        
        contacts = Contact.objects.filter(
            id__in=contact_ids,
            organization=self.request.user.organization
        )
        
        serializer = ContactSerializer(contacts, many=True)
        return Response(serializer.data)


class ContactViewSet(BaseOrganizationViewSet):
    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_subscribed', 'language']
    search_fields = ['email', 'first_name', 'last_name']
    pagination_class = DefaultPagination
    
    def create(self, request, *args, **kwargs):
        """Create a new contact with subscription limit check"""
        if not request.user.organization:
            return Response({
                'error': 'No organization found'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check contact limit
        limit_check = check_usage_limit(request.user.organization, 'contacts')
        if not limit_check['allowed']:
            return Response({
                'error': limit_check['reason'],
                'code': 'CONTACT_LIMIT_REACHED',
                'current': limit_check['current'],
                'limit': limit_check['limit'],
                'upgrade_url': '/subscription/upgrade'
            }, status=status.HTTP_403_FORBIDDEN)
        
        response = super().create(request, *args, **kwargs)
        
        # Update usage tracking
        if response.status_code == 201:
            update_usage_tracking(request.user.organization, 'contacts')
        
        return response

    @action(detail=False, methods=['post'])
    def import_csv(self, request):
        """Import contacts from CSV file with subscription limit check"""
        # Check contact limit before import
        limit_check = check_usage_limit(request.user.organization, 'contacts')
        if not limit_check['allowed']:
            return Response({
                'error': limit_check['reason'],
                'code': 'CONTACT_LIMIT_REACHED',
                'current': limit_check['current'],
                'limit': limit_check['limit'],
                'upgrade_url': '/subscription/upgrade'
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = ContactImportSerializer(data=request.data)
        if serializer.is_valid():
            file = serializer.validated_data['file']
            
            # Read file content
            if file.name.lower().endswith('.csv'):
                content = file.read().decode('utf-8')
            elif file.name.lower().endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                
                # Convert Excel to CSV format
                output = io.StringIO()
                writer = csv.writer(output)
                
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(row)
                
                content = output.getvalue()
            
            # Start background task
            task = import_contacts_from_csv.delay(
                self.request.user.organization.id,
                content,
                self.request.user.id
            )
            
            # Update usage tracking
            update_usage_tracking(request.user.organization, 'contacts')
            
            return Response({
                'message': 'Import started',
                'task_id': task.id
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """Bulk update contacts"""
        contact_ids = request.data.get('contact_ids', [])
        update_data = request.data.get('data', {})
        
        contacts = Contact.objects.filter(
            id__in=contact_ids,
            organization=self.request.user.organization
        )
        
        updated_count = contacts.update(**update_data)
        
        return Response({
            'message': f'Updated {updated_count} contacts'
        })

    @action(detail=False, methods=['get'])
    def export_csv(self, request):
        """Export contacts as CSV or Excel file"""
        from django.http import HttpResponse
        import csv
        import io
        import openpyxl
        
        # Get format parameter (csv or xlsx)
        export_format = request.query_params.get('format', 'csv').lower()
        
        # Get filtered contacts based on query parameters
        queryset = self.get_queryset()
        
        # Apply filters similar to the list view
        search = request.query_params.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(email__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search)
            )
        
        is_subscribed = request.query_params.get('is_subscribed', '')
        if is_subscribed in ['true', 'false']:
            queryset = queryset.filter(is_subscribed=is_subscribed == 'true')
        
        language = request.query_params.get('language', '')
        if language:
            queryset = queryset.filter(language=language)
        
        # Helper function to sanitize fields to prevent injection
        def sanitize_field(value):
            """Sanitize field to prevent CSV/Excel injection attacks"""
            if not value:
                return ''
            value_str = str(value)
            # Strip leading whitespace and tabs, then check for dangerous characters
            stripped_value = value_str.lstrip(' \t')
            if stripped_value.startswith(('=', '+', '-', '@')):
                return "'" + value_str
            return value_str
        
        # Prepare data rows
        headers = [
            'email', 'first_name', 'last_name', 'phone', 
            'language', 'is_subscribed', 'created_at', 'updated_at'
        ]
        
        rows = []
        for contact in queryset.order_by('email'):
            rows.append([
                sanitize_field(contact.email),
                sanitize_field(contact.first_name or ''),
                sanitize_field(contact.last_name or ''),
                sanitize_field(contact.phone or ''),
                sanitize_field(contact.language or 'en'),
                'Yes' if contact.is_subscribed else 'No',
                contact.created_at.strftime('%Y-%m-%d %H:%M:%S') if contact.created_at else '',
                contact.updated_at.strftime('%Y-%m-%d %H:%M:%S') if contact.updated_at else ''
            ])
        
        if export_format == 'xlsx':
            # Create Excel file
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Contacts"
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_idx, value=header)
            
            # Write data rows
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, cell_value in enumerate(row_data, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            
            # Create response
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="contacts_export.xlsx"'
            
        else:
            # Create CSV response (default)
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="contacts_export.csv"'
            
            writer = csv.writer(response)
            writer.writerow(headers)
            writer.writerows(rows)
        
        return response


class EmailTemplateViewSet(BaseOrganizationViewSet):
    queryset = EmailTemplate.objects.all()
    serializer_class = EmailTemplateSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['category', 'is_default']
    search_fields = ['name', 'subject']
    
    def create(self, request, *args, **kwargs):
        """Create a new template with limit check"""
        if not request.user.organization:
            return Response({
                'error': 'No organization found'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check template limit
        limit_check = check_usage_limit(request.user.organization, 'templates')
        if not limit_check['allowed']:
            return Response({
                'error': limit_check['reason'],
                'code': 'TEMPLATE_LIMIT_REACHED',
                'current': limit_check['current'],
                'limit': limit_check['limit'],
                'upgrade_url': '/subscription/upgrade'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check custom templates feature for advanced templates
        if request.data.get('is_custom', False):
            if not check_feature_available(request.user.organization, 'has_custom_templates'):
                return Response({
                    'error': 'Custom templates are not available in your current plan',
                    'code': 'FEATURE_NOT_AVAILABLE',
                    'feature': 'has_custom_templates',
                    'upgrade_url': '/subscription/upgrade'
                }, status=status.HTTP_403_FORBIDDEN)
        
        response = super().create(request, *args, **kwargs)
        
        # Update usage tracking if successful
        if response.status_code == 201:
            update_usage_tracking(request.user.organization, 'templates')
        
        return response

    def list(self, request, *args, **kwargs):
        """List templates and seed defaults if none exist"""
        queryset = self.get_queryset()
        
        # If no templates exist, seed default templates
        if not queryset.exists():
            self.seed_default_templates()
            queryset = self.get_queryset()
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def seed_default_templates(self):
        """Seed default email templates"""
        if not self.request.user.organization:
            return
            
        default_templates = [
            {
                'name': 'Welcome Email',
                'subject': 'Welcome to {{organizationName}}! 🎉',
                'category': 'marketing',
                'is_default': True,
                'html_content': '''<div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h1 style="color: #333; text-align: center;">Welcome to Our Community!</h1>
                    <p>Hi {{firstName}},</p>
                    <p>We're thrilled to have you join us! Your account has been successfully created.</p>
                    <div style="background-color: #f8f9fa; padding: 20px; border-radius: 8px; margin: 20px 0;">
                        <h3 style="color: #495057; margin-top: 0;">What's Next?</h3>
                        <ul><li>Complete your profile</li><li>Explore features</li><li>Connect with others</li></ul>
                    </div>
                    <p>Best regards,<br>The {{organizationName}} Team</p>
                </div>''',
                'text_content': '''Hi {{firstName}},

Welcome to {{organizationName}}! We're thrilled to have you join us.

What's Next?
- Complete your profile
- Explore our features
- Connect with other members

Best regards,
The {{organizationName}} Team'''
            },
            {
                'name': 'Order Confirmation',
                'subject': 'Order Confirmation #{{orderNumber}}',
                'category': 'transactional',
                'is_default': True,
                'html_content': '''<div style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
                    <h1 style="color: #28a745; text-align: center;">Order Confirmed! ✅</h1>
                    <p>Hi {{firstName}},</p>
                    <p>Thank you for your order! We've received your payment and your order is being processed.</p>
                    <div style="background-color: #e8f5e8; padding: 20px; border-radius: 8px; margin: 20px 0;">
                        <h3>Order Details</h3>
                        <p><strong>Order Number:</strong> #{{orderNumber}}</p>
                        <p><strong>Total Amount:</strong> ${{orderTotal}}</p>
                    </div>
                    <p>Thank you for choosing {{organizationName}}!</p>
                </div>''',
                'text_content': '''Hi {{firstName}},

Thank you for your order! We've received your payment.

Order Details:
- Order Number: #{{orderNumber}}
- Total Amount: ${{orderTotal}}

Thank you for choosing {{organizationName}}!'''
            }
        ]
        
        for template_data in default_templates:
            EmailTemplate.objects.get_or_create(
                organization=self.request.user.organization,
                name=template_data['name'],
                defaults=template_data
            )


class CampaignViewSet(BaseOrganizationViewSet):
    queryset = Campaign.objects.all()
    serializer_class = CampaignSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status']
    search_fields = ['name', 'subject']
    ordering_fields = ['created_at', 'sent_at', 'name']
    ordering = ['-created_at']
    pagination_class = DefaultPagination
    
    def create(self, request, *args, **kwargs):
        """Create a new campaign with subscription limit check"""
        import logging
        logger = logging.getLogger('django')
        
        logger.error(f"=== CAMPAIGN CREATION DEBUG ===")
        logger.error(f"User: {request.user.id} - {request.user.email}")
        logger.error(f"User organization: {request.user.organization}")  
        logger.error(f"Request data: {request.data}")
        
        if not request.user.organization:
            logger.error("No organization found for user")
            return Response({
                'error': 'No organization found'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check campaign limit
        limit_check = check_usage_limit(request.user.organization, 'campaigns')
        if not limit_check['allowed']:
            logger.error(f"Campaign limit reached: {limit_check}")
            return Response({
                'error': limit_check['reason'],
                'code': 'CAMPAIGN_LIMIT_REACHED',
                'current': limit_check['current'],
                'limit': limit_check['limit'],
                'upgrade_url': '/subscription/upgrade'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Check A/B testing feature if it's an A/B test campaign
        if request.data.get('is_ab_test', False):
            if not check_feature_available(request.user.organization, 'has_ab_testing'):
                logger.error("A/B testing not available")
                return Response({
                    'error': 'A/B testing is not available in your current plan',
                    'code': 'FEATURE_NOT_AVAILABLE',
                    'feature': 'has_ab_testing',
                    'upgrade_url': '/subscription/upgrade'
                }, status=status.HTTP_403_FORBIDDEN)
        
        # Create the campaign
        try:
            logger.error("About to call super().create()")
            response = super().create(request, *args, **kwargs)
            logger.error(f"Super create response: {response.status_code}")
            if hasattr(response, 'data'):
                logger.error(f"Response data: {response.data}")
            
            # Update usage tracking if successful
            if response.status_code == 201:
                update_usage_tracking(request.user.organization, 'campaigns')
                logger.error("Campaign created successfully!")
            
            return response
        except Exception as e:
            logger.error(f"Exception in campaign creation: {str(e)}")
            logger.error(f"Exception type: {type(e)}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")
            return Response({
                'error': f'Campaign creation failed: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    def perform_create(self, serializer):
        serializer.save(
            organization=self.request.user.organization,
            created_by=self.request.user
        )

    @action(detail=True, methods=['post'])
    def send(self, request, pk=None):
        """Send campaign emails with subscription and email limit checks"""
        campaign = self.get_object()
        
        # Check subscription is active
        if not check_subscription_active(request.user.organization):
            return Response({
                'error': 'Active subscription required to send campaigns',
                'code': 'SUBSCRIPTION_REQUIRED',
                'upgrade_url': '/subscription/plans'
            }, status=status.HTTP_402_PAYMENT_REQUIRED)
        
        if campaign.status != 'draft':
            return Response(
                {'error': 'Campaign must be in draft status to send'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = CampaignSendSerializer(data=request.data)
        if serializer.is_valid():
            data = serializer.validated_data
            
            # Get recipient contacts
            contacts = Contact.objects.filter(
                organization=self.request.user.organization,
                is_subscribed=True
            )
            
            if data.get('send_all'):
                # Send to all subscribed contacts
                pass
            elif data.get('contact_ids'):
                contacts = contacts.filter(id__in=data['contact_ids'])
            elif data.get('group_ids'):
                contacts = contacts.filter(
                    group_memberships__group_id__in=data['group_ids']
                ).distinct()
            
            # Check monthly email limit before sending
            recipient_count = contacts.count()
            usage = get_current_usage(request.user.organization)
            limit_check = check_usage_limit(request.user.organization, 'emails')
            
            # Check if sending would exceed limit
            if usage['emails_sent'] + recipient_count > limit_check.get('limit', 0):
                return Response({
                    'error': f'Sending this campaign would exceed your monthly email limit',
                    'code': 'EMAIL_LIMIT_EXCEEDED',
                    'current': usage['emails_sent'],
                    'would_send': recipient_count,
                    'limit': limit_check.get('limit', 0),
                    'upgrade_url': '/subscription/upgrade'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Create campaign recipients
            recipients = []
            for contact in contacts:
                recipients.append(
                    CampaignRecipient(
                        campaign=campaign,
                        contact=contact,
                        status='pending'
                    )
                )
            
            CampaignRecipient.objects.bulk_create(recipients, ignore_conflicts=True)
            
            # Update campaign statistics
            campaign.total_recipients = len(recipients)
            campaign.save()
            
            # Start sending emails asynchronously
            send_campaign_emails.delay(campaign.id)
            
            # Update usage tracking for emails
            update_usage_tracking(request.user.organization, 'emails', len(recipients))
            
            return Response({
                'message': f'Campaign queued for sending to {len(recipients)} recipients',
                'recipients': len(recipients),
                'usage': {
                    'emails_sent_before': usage['emails_sent'],
                    'emails_sent_after': usage['emails_sent'] + len(recipients),
                    'monthly_limit': limit_check.get('limit', 0)
                }
            })
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Duplicate a campaign with limit check"""
        # Check campaign limit before duplication
        limit_check = check_usage_limit(request.user.organization, 'campaigns')
        if not limit_check['allowed']:
            return Response({
                'error': limit_check['reason'],
                'code': 'CAMPAIGN_LIMIT_REACHED',
                'current': limit_check['current'],
                'limit': limit_check['limit'],
                'upgrade_url': '/subscription/upgrade'
            }, status=status.HTTP_403_FORBIDDEN)
        
        original_campaign = self.get_object()
        
        # Create copy
        new_campaign = Campaign(
            organization=original_campaign.organization,
            name=f"{original_campaign.name} (Copy)",
            subject=original_campaign.subject,
            from_email=original_campaign.from_email,
            from_name=original_campaign.from_name,
            reply_to_email=original_campaign.reply_to_email,
            html_content=original_campaign.html_content,
            text_content=original_campaign.text_content,
            created_by=self.request.user,
            status='draft'
        )
        new_campaign.save()
        
        # Update usage tracking
        update_usage_tracking(request.user.organization, 'campaigns')
        
        serializer = self.get_serializer(new_campaign)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class AnalyticsEventViewSet(BaseOrganizationViewSet):
    queryset = AnalyticsEvent.objects.all()
    serializer_class = AnalyticsEventSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['event_type', 'campaign']
    ordering = ['-created_at']
    
    @action(detail=False, methods=['get'])
    @requires_plan_feature('has_advanced_analytics')
    def advanced_reports(self, request):
        """Get advanced analytics reports - Pro/Premium feature"""
        # Advanced analytics logic here
        org = request.user.organization
        
        # Get advanced metrics
        from django.db.models import Avg, Max, Min
        campaign_stats = Campaign.objects.filter(organization=org).aggregate(
            avg_open_rate=Avg('open_rate'),
            avg_click_rate=Avg('click_rate'),
            max_open_rate=Max('open_rate'),
            min_open_rate=Min('open_rate')
        )
        
        return Response({
            'message': 'Advanced analytics available',
            'reports': campaign_stats
        })
    
    @action(detail=False, methods=['get'])
    def basic_reports(self, request):
        """Get basic analytics reports - available to all plans"""
        # Basic analytics logic here
        org = request.user.organization
        
        # Get basic counts
        total_sent = Campaign.objects.filter(organization=org).aggregate(
            total=Sum('total_sent')
        )['total'] or 0
        
        return Response({
            'message': 'Basic analytics',
            'total_sent': total_sent
        })


class DashboardViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get dashboard statistics"""
        if not request.user.organization:
            return Response({'error': 'No organization found'}, status=status.HTTP_400_BAD_REQUEST)
        
        org = request.user.organization
        
        # Calculate statistics
        total_contacts = Contact.objects.filter(organization=org).count()
        active_campaigns = Campaign.objects.filter(
            organization=org,
            status__in=['sending', 'scheduled']
        ).count()
        
        # Aggregate campaign statistics
        campaign_stats = Campaign.objects.filter(organization=org).aggregate(
            total_sent=Sum('total_sent'),
            total_opened=Sum('total_opened'),
            total_clicked=Sum('total_clicked')
        )
        
        total_sent = campaign_stats['total_sent'] or 0
        total_opened = campaign_stats['total_opened'] or 0
        total_clicked = campaign_stats['total_clicked'] or 0
        
        open_rate = (total_opened / total_sent * 100) if total_sent > 0 else 0
        click_rate = (total_clicked / total_sent * 100) if total_sent > 0 else 0
        
        stats_data = {
            'total_contacts': total_contacts,
            'active_campaigns': active_campaigns,
            'total_sent': total_sent,
            'total_opened': total_opened,
            'total_clicked': total_clicked,
            'open_rate': round(open_rate, 2),
            'click_rate': round(click_rate, 2)
        }
        
        serializer = DashboardStatsSerializer(stats_data)
        return Response(serializer.data)


class CardViewSet(BaseOrganizationViewSet):
    """ViewSet for managing cards"""
    queryset = Card.objects.all()
    serializer_class = CardSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """Get cards for the current organization"""
        if not self.request.user.organization:
            return Card.objects.none()
        return Card.objects.filter(
            organization=self.request.user.organization
        ).order_by('-is_default', '-created_at')

    def get_serializer_class(self):
        """Use different serializers for different actions"""
        if self.action == 'create':
            return CardCreateSerializer
        elif self.action == 'update' or self.action == 'partial_update':
            return CardUpdateSerializer
        return CardSerializer

    def create(self, request, *args, **kwargs):
        """Create a new card with direct card details"""        
        if not request.user.organization:
            return Response({
                'error': 'User not associated with organization'
            }, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        org = request.user.organization
        
        try:
            # Extract safe card details from validated data
            validated_data = serializer.validated_data
            
            # Create card record with safe details only (PCI compliant)
            card = Card.objects.create(
                organization=org,
                cardholder_name=validated_data.get('cardholder_name'),
                last4=validated_data.get('last4'),  # Already validated to be 4 digits
                brand=validated_data.get('brand'),  # Already validated brand
                exp_month=validated_data.get('exp_month'),
                exp_year=validated_data.get('exp_year'),
                is_default=validated_data.get('set_as_default', True)
            )

            return Response(
                CardSerializer(card).data,
                status=status.HTTP_201_CREATED
            )

        except Exception as e:
            return Response({
                'error': f'Error creating card: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



    def destroy(self, request, pk=None):
        """Delete a card"""
        try:
            card = self.get_object()
            
            # If this was the default card, set another card as default
            if card.is_default:
                org = request.user.organization
                other_card = Card.objects.filter(
                    organization=org
                ).exclude(id=card.id).first()
                
                if other_card:
                    other_card.is_default = True
                    other_card.save()
            
            # Delete the card
            card.delete()
            
            return Response({'message': 'Card deleted successfully'})
            
        except Exception as e:
            return Response({
                'error': f'Error deleting card: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def default(self, request):
        """Get the default card for the organization"""
        if not request.user.organization:
            return Response({
                'error': 'User not associated with organization'
            }, status=status.HTTP_400_BAD_REQUEST)

        default_card = Card.objects.filter(
            organization=request.user.organization,
            is_default=True
        ).first()

        if not default_card:
            return Response({
                'message': 'No default card found'
            }, status=status.HTTP_404_NOT_FOUND)

        return Response(CardSerializer(default_card).data)